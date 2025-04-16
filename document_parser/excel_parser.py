import json, os, re, requests
import pandas as pd

from openpyxl import load_workbook
from pydantic import BaseModel
from utils.logging import get_logger
logger = get_logger("excel_parser")

def get_topk_rows_from_file(filepath, k=5):
    res = {}
 
    # # .xlsx
    # wb = load_workbook(filepath)
    # list_sheets = wb.sheetnames

    # adapt to .xlsx / .xls / .csv
    file_extension = os.path.splitext(filepath)[1].lower()
    if file_extension == ".xlsx" or file_extension == ".xls":
        file_in = pd.ExcelFile(filepath)
        list_sheets = file_in.sheet_names
    
        for s in list_sheets:
            df = pd.read_excel(filepath, sheet_name=s, header=None, nrows=k)
            res[s] = df.values.tolist()
    elif file_extension == ".csv":
        df = pd.read_csv(filepath, header=None, nrows=k)
        res['Sheet1'] = df.values.tolist()
    else:
        logger.error(f"不支持的文件格式: {filepath}")

    return res

def get_clean_row(row:list):
    cleaned_row =  [x for x in row if not pd.isna(x)]
    return cleaned_row

def json_extractor(text:str):
    json_pattern = r"```json\n(.*?)\n```"
    match = re.search(json_pattern, text, re.DOTALL)
    if match:
        # logger.debug(f"Found JSON pattern: {match}")
        json_str = match.group(0).replace("```json", "").replace("```", "").strip()
        json_str = json_str.replace("True", "true").replace("False", "false")
        try:
            json_data = json.loads(json_str)
            return json_data
        except json.JSONDecodeError as e:
            logger.error(f"JSON字符串:{json_str}\nJSON解析错误: {e}")
            return None
    else:
        logger.error(f"未找到JSON数据")
        return None 

class HeaderDetector(BaseModel):
    """
    表头结构解析器
    - 是否为表头行（是/否）
    - 是否属于复合表头（是/否）
    - 置信度评分（0-100）
    """
    is_header: bool
    is_compound_header: bool
    confidence: int

def detect_headers_llm(excel_data, model="qwen2.5:latest", k=5):
    """
    使用 LLM 解析 Excel 数据内容，获取表头信息
    :param model: LLM 模型
    :param excel_data: Excel 数据前5行内容
    :return: 表头信息
    """
    
    # 调用 LLM 进行解析
    base_url = os.getenv("LLM_HOST")
    # logger.debug(f"detecting headers via llm host: {base_url}")
    excel_structure = {}

    for sheet_name, sheet_data in excel_data.items():
        excel_structure[sheet_name] = []
        for idx, row in enumerate(sheet_data):
            cleaned_row = get_clean_row(row)
            if len(cleaned_row) <= 1:
                continue
            prev_row = [x for x in sheet_data[idx-1]] if idx > 0 else None
            next_row = [x for x in sheet_data[idx+1]] if idx < len(sheet_data)-1 else None
            print
            sys_prompt = "你是表格结构分析专家，擅长识别带占位符的复合表头。"
            user_prompt = f"""
                请分析以下表格行是否为可能的表头行或复合表头组成部分：
                - 表格名称：{sheet_name}
                - 行号：{idx + 1}
                - 行内容：{row}
                - 前一行内容：{prev_row}
                - 后一行内容：{next_row}
        
                注意事项：
                1. **空值（NaN）**表示占位空格，可能用于合并单元格或层次结构。
                2. **连续NaN**表示跨列合并单元格（如某行连续多个NaN表示该行被合并到左侧单元格）。
                3. **复合表头特征**：
                - 主表头可能占据多行，覆盖多个列，同时应包含具体列名。
                - 子表头通常在主表头下方对齐，与主表头形成父子关系。
                - 主表头和子表头通常是在连续的行中，不会存在不连续的行中
                4. **标题行不是表头**
                5. **数据行**通常包含具体数值或名称、类别标签等，而非列名。

                判断依据：
                1. **是否为表头行**：该行是否包含列名（如"序号"、"项目名称"、"年度"等）。
                2. **是否为复合表头**：
                - 该行是否与前/后行形成层次结构（如主表头后接子表头）。
                - 是否存在跨列合并单元格（连续NaN）。
                - 是否存在多行标题模式（如第一行概括，后续行细分）。
                
                按如下json格式返回结果，无需额外解释：
                ```json
                {{
                    "is_header": boolean, 是否为表头行（True/False）,
                    "is_compound_header": boolean, 是否属于复合表头（True/False）,
                    "confidence": int, 置信度评分（0-100）                
                }}
                ```
                
            """
            messages = [
                {"role":"system", "content":sys_prompt},
                {"role":"user", "content":user_prompt}
                ]
            data = {
                "model": model,
                "messages": messages,
                "format": HeaderDetector.model_json_schema(),
                "temperature": 0.1
            }
            logger.info(f"request data: {data}")
            response = requests.post(f"{base_url}/v1/chat/completions", json=data).json()
            if response.get("error"):
                logger.error(f"LLM 解析错误: {response['error']['message']}")
                return None
            if response.get("choices"):
                content = response["choices"][0]["message"]["content"]
                logger.info(f"LLM 解析结果: {content}")
                res = json_extractor(content)
                if res:
                    logger.info(f"当前行{idx}: {row}\n解析结果: {res}")
                    # print(res["is_header"])
                    # if res["is_header"] and res["is_compound_header"] and res["confidence"] > 80:
                    if res["is_header"]:
                        excel_structure[sheet_name].append(idx)
                    elif len(excel_structure[sheet_name]) > 0:
                        break
                    else: 
                        continue
    logger.info(f"table structure: {excel_structure}")
    return excel_structure

if __name__ == "__main__":
    # testfile = "/Users/zhourong/Downloads/附件：2024年度基础研究面上项目拟资助项目清单.xlsx"
    # testfile = "/Users/zhourong/Desktop/temp/3. 2020-2023年深圳市卫生事业发展情况数据.xlsx"
    # testfile = "tmp/2020-2023年深圳市卫生事业发展情况数据.xlsx"
    testfile = "/home/ubuntu/data/excel/一句话查询用户场景（全）.xlsx"
    topk_rows = get_topk_rows_from_file(testfile)
    # print(f"topk_rows: {topk_rows}")
    excel_headers = detect_headers_llm(topk_rows, model="qwen2.5:72b")
    print(f"excel_headers: {excel_headers}")
    
